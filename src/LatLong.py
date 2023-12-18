import openpyxl, requests, folium, sys

excel = openpyxl.load_workbook("C:\\Users\\Scott\\Documents\\TestData.xlsx")
sheet = excel.active

CityState = {}

# Gathering each unique City and State
for cell in range(2,sheet.max_row+1):
    
    if sheet["D"+str(cell)].value not in CityState:
        CityState[sheet["D"+str(cell)].value] = [sheet["E"+str(cell)].value, sheet["M"+str(cell)].value, 1]
        
    if sheet["D"+str(cell)].value in CityState:
        CityState[sheet["D"+str(cell)].value][1] += sheet["M"+str(cell)].value
        CityState[sheet["D"+str(cell)].value][2] += 1

# Google Maps API key
p = open('C:\\Users\\Scott\\Documents\\apikey.txt', "r")
api_key = p.read()

excel.save("C:\\Users\\Scott\\Documents\\TestData.xlsx")
excel.close()

# Write next gathered information to new workbook
workbook = openpyxl.Workbook() 
sheet = workbook.active 

cell = 2

# Reformat and get geolocation using api key then place in new workbook
for i in CityState:
    i.replace(" ","+")
    response = requests.get(f"https://maps.googleapis.com/maps/api/geocode/json?address=+{i},+{CityState[i][0]}&key={api_key}")

    resp_json_payload = response.json()
    
    sheet["A"+str(cell)].value = i
    sheet["B"+str(cell)].value = CityState[i][0]
    sheet["C"+str(cell)].value = CityState[i][1]
    sheet["D"+str(cell)].value = CityState[i][2]
    sheet["E"+str(cell)].value = str(resp_json_payload['results'][0]['geometry']['location'])
    cell += 1
    
p.close()      

# strip the json response
for cell in range(2,sheet.max_row+1):
    clean = sheet["E"+str(cell)].value.translate({ord(i): None for i in "'latng{}: "})
    LatLong = clean.split(",")
    sheet["F"+str(cell)].value = LatLong[0]
    sheet["G"+str(cell)].value = LatLong[1]
    

workbook.save("C:\\Users\\Scott\\Documents\\DistanceInfo.xlsx")
workbook.close()

# Define functions to give values to map based on percentile total gallons or shipments each city received
def colorFun(cell):
    totalGallon = sheet["C"+str(cell)].value
    if totalGallon <= 36:
        return "yellow"
    if totalGallon <= 298.469:
        return "green"
    if totalGallon <= 2025.19:
        return "orange"
    if totalGallon >= 2025.19:
        return "red"
    else:
        return "black"
def weightFun(cell):
    totalShipments = sheet["D"+str(cell)].value
    if totalShipments <= 3:
        return 0.75
    if totalShipments <= 5:
        return 1
    if totalShipments <= 18:
        return 1.25
    if totalShipments >= 18:
        return 1.5
    else:
        return 10
my_map4 = folium.Map(location = [32.5251516, -93.7501789], tiles="Cartodb Positron",
										zoom_start = 6) 

folium.Marker([32.5251516, -93.7501789], 
			tooltip = 'Shreveport').add_to(my_map4) 
sys.setrecursionlimit(10000000)
excel = openpyxl.load_workbook("C:\\Users\\Scott\\Documents\\DistanceInfo.xlsx")
sheet = excel.active
lat = []
long = []
Names = []
Coord = []
for cell in range(2,sheet.max_row+1):
    lat.append(float(sheet["F"+str(cell)].value))
    long.append(float(sheet["G"+str(cell)].value))
    Names.append(sheet["A"+str(cell)].value)
    
num = 0
for i in lat:
    folium.Marker([i, long[num]], tooltip = Names[num]).add_to(my_map4)  
    Coord.append([i,long[num]])
    folium.PolyLine(locations = [Coord[num], (32.5251516, -93.7501789)], color=colorFun(num+2),
    				 weight=weightFun(num+2)).add_to(my_map4) 
    num +=1

excel.close()
my_map4.save("my_map7.html") 
